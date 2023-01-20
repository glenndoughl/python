from openpyxl import Workbook, load_workbook
import csv
import time
from datetime import datetime
from dateutil.relativedelta import relativedelta

def main():

    day_offset = 1#input("Day Offset: ")
    yyyymmdd_min1d = (datetime.now()+relativedelta(days=day_offset)).strftime("%Y%m%d") #get date in yyyymmdd-1day (ie: 20221018)
    yyyy = datetime.now().strftime("%Y") #get year in YYYY format (ie: 2022)

    if int(datetime.now().strftime("%d")) > 25:
        month_offset = 1
    else:
        month_offset = 0

    mm_Mmmm = (datetime.now()+relativedelta(months=month_offset)).strftime("%m %B")  #get current month in MM_MMMM format (ie: 10 October)
    mmmm = (datetime.now()+relativedelta(months=month_offset)).strftime("%m %B")  #get month in MMMM (ie: October)
    yyyy = datetime.now().strftime("%Y") #get year in YYYY format (ie: 2022)

    CIP = input("CIP new DGO: ")
    BPGC = input("BPGC new DGO: ")
    OSPGC = input("OSPGC new DGO: ")
    values = [CIP, BPGC, OSPGC]

    files_conf = {
    
    "file1":{
        "path":rf"C:\Users\aspan.clr\AC Energy, Inc\Internal Document Library - 04 Trading\Trading Operations\10. ST",
        "sheetname":"Price vs. RTD",
        "cells":["S3","T3","U3"],
        "newvalue": values
        },
    "file2":{
        "path":rf"C:\Users\aspan.clr\AC Energy, Inc\Internal Document Library - 04 Trading\Trading Operations\10. ST\GAMBIT\{yyyy}\{mm_Mmmm}\{yyyy}",
        "sheetname":"Data",
        "cells":["B2","B3","B4"],
        "newvalue": values
        },
    "file3":{
        "path":rf"C:\Users\aspan.clr\AC Energy, Inc\Internal Document Library - 04 Trading\Trading Operations\03.PM\BCQ\EWDO\MQ CONVERTER\{yyyymmdd_min1d}",
        "sheetname":"main",
        "cells":["C3","C4","C2"],
        "newvalue": values
        },
    "file4":{
        "path":rf"C:\Users\aspan.clr\AC Energy, Inc\Internal Document Library - 04 Trading\Trading Operations\01.AM\ASPA\AM2 Cheat Sheet",
        "sheetname":"Forecast GP",
        "cells":["C1","C2","C3"],
        "newvalue": values
        },
    }

    for conf in files_conf:
        
        FilePath = files_conf[conf]["path"]
        SheetName =  files_conf[conf]["sheetname"]
        CellNumbers =  files_conf[conf]["cells"]
        New_Value = files_conf[conf]["newvalue"]

        print(f"""
        EDITING:
        File: {FilePath}
        Sheet: {SheetName}
        Plants: [CIP,BPGC,OSPGC]
        Cells: {CellNumbers}
        New Values: {New_Value}""")


        wb = load_workbook(rf"{FilePath}", keep_vba=True)
        ws = wb[SheetName]
        for i in range(3):
            ws[CellNumbers[i]] = New_Value[i]
            print(f"{CellNumbers[i]} = {New_Value[i]}")
        wb.save(FilePath)

        print(f"""DONE - Saved to: {FilePath}
    **************************************************************************""")

if __name__=="__main__":
    main()